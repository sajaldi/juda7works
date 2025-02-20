import openpyxl
from django.http import HttpResponse
from django.apps import apps
from django.core.files.storage import default_storage





def export_to_excel(model_name):
    """Exporta los datos de un modelo a un archivo Excel, incluyendo el ID."""
    model = apps.get_model(app_label='cmms', model_name=model_name)
    objects = model.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name

    # Crear encabezados, incluyendo el ID
    headers = [field.name for field in model._meta.fields]
    ws.append(headers)

    for obj in objects:
        # Agregar los datos del modelo, incluyendo el ID
        row = []
        for field in model._meta.fields:
            value = getattr(obj, field.name)

            # Si el campo es una relación, extraer el valor de la relación
            if hasattr(value, '__str__'):  # Verifica si el valor tiene un método __str__
                row.append(str(value))  # Usamos str() para obtener una representación legible
            else:
                row.append(value)  # Si no es una relación, agregamos el valor directamente

        ws.append(row)

    # Guardar el archivo Excel en memoria
    file_path = f'{model_name}_export.xlsx'
    wb.save(file_path)
    return file_path


from django.core.exceptions import ObjectDoesNotExist

import openpyxl
from django.apps import apps


from django.apps import apps



from django.db import models
from django.apps import apps
from django.core.exceptions import ObjectDoesNotExist
import openpyxl

def import_from_excel(model_name, file_path):
    """
    Importa datos de un archivo Excel a un modelo de Django, con lógica de actualización.
    Para campos ForeignKey, se intenta convertir un valor de cadena (por ejemplo, "ThyssenKrup")
    en la instancia correspondiente, comparando con la representación en cadena de cada instancia.
    """
    model = apps.get_model(app_label='cmms', model_name=model_name)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Los encabezados del Excel se toman de la primera fila.
    headers = [cell.value for cell in ws[1]]
    # Se crea un diccionario con los campos del modelo.
    model_fields = {field.name: field for field in model._meta.fields}

    # Se verifica que todos los encabezados estén en los campos del modelo.
    if not set(headers).issubset(model_fields.keys()):
        raise ValueError("Los encabezados del archivo no coinciden con los campos del modelo.")

    updated_objects = []
    new_objects = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        obj_id = data.get('id')

        # Procesamos los campos ForeignKey de forma general.
        for field_name, field in model_fields.items():
            if isinstance(field, models.ForeignKey):
                raw_value = data.get(field_name)
                if isinstance(raw_value, str) and raw_value.strip():
                    related_model = field.related_model
                    found_instance = None
                    # Iteramos sobre las instancias del modelo relacionado para encontrar una coincidencia
                    for instance in related_model.objects.all():
                        # Se compara la representación en cadena, eliminando espacios en blanco.
                        if str(instance).strip() == raw_value.strip():
                            found_instance = instance
                            break
                    data[field_name] = found_instance  # Si no se encuentra, quedará como None.
        
        # Si se proporciona un ID, se intenta actualizar el objeto existente.
        if obj_id:
            try:
                obj = model.objects.get(id=obj_id)
                for field, value in data.items():
                    if field != 'id':  # No se actualiza el ID.
                        setattr(obj, field, value)
                updated_objects.append(obj)
            except ObjectDoesNotExist:
                new_objects.append(model(**data))
        else:
            new_objects.append(model(**data))
    
    if new_objects:
        model.objects.bulk_create(new_objects)
    for obj in updated_objects:
        obj.save()

    return f"Se importaron {len(new_objects)} registros nuevos y se actualizaron {len(updated_objects)} registros existentes en {model_name}."
