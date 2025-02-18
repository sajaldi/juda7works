from django.contrib import admin, messages
from django.shortcuts import redirect, render
from django.http import HttpResponse
import openpyxl

# Asegúrate de importar el modelo correctamente
from ...models import Sistema


def importar_sistemas_excel(request):
    """
    Vista para importar sistemas desde un archivo Excel, actualizando existentes por ID o creando nuevos.
    """
    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Error al abrir el archivo: {str(e)}")
            return redirect("importar_sistemas")  # Asegúrate de tener una URL llamada 'importar_sistemas'

        errores = []
        # Iterar a partir de la fila 2 para saltar los encabezados
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            print(f"Fila {i}: {row}")  # Depuración: ver qué datos se están leyendo

            # Validar que la fila tenga exactamente 3 columnas (ID, Nombre, Sistema Principal)
            if row is None or len(row) != 3:
                errores.append(f"Fila {i}: Número incorrecto de columnas ({len(row) if row else 0}). Se esperaban 3 (ID, Nombre, Sistema Principal).")
                continue

            sistema_id, sistema_nombre, sistema_principal_nombre = row

            # Validar que el valor de 'Nombre' no sea None (el sistema principal puede ser None)
            if sistema_nombre is None:
                errores.append(f"Fila {i}: La celda 'Nombre' está vacía.")
                continue

            # Buscar el sistema principal si se proporciona un nombre
            sistema_principal = None
            if sistema_principal_nombre:
                sistema_principal = Sistema.objects.filter(nombre=sistema_principal_nombre).first()
                if not sistema_principal and sistema_principal_nombre.strip():
                    errores.append(f"Fila {i}: El sistema principal '{sistema_principal_nombre}' no existe.")
                    continue

            # Intentar actualizar o crear el sistema basado en el ID
            try:
                sistema_obj, created = Sistema.objects.update_or_create(
                    id=sistema_id,  # Usamos el ID del archivo para buscar o crear
                    defaults={
                        'nombre': sistema_nombre,
                        'principal': sistema_principal
                    }
                )
                if created:
                    print(f"Fila {i}: Sistema con ID '{sistema_id}' creado.")
                else:
                    print(f"Fila {i}: Sistema con ID '{sistema_id}' actualizado.")
            except Exception as e:
                errores.append(f"Fila {i}: Error al guardar/actualizar en la base de datos para ID '{sistema_id}' ({str(e)}).")

        # Mostrar mensajes al usuario según se hayan producido errores o no
        if errores:
            messages.error(request, "Errores en la importación:\n" + "\n".join(errores))
        else:
            messages.success(request, "Importación exitosa.")

        return redirect("importar_sistemas")  # Redirige a la vista de importación

    return render(request, "cmms/importar_sistemas.html")  # Asegúrate de tener este template


def exportar_sistemas_seleccionados_excel(modeladmin, request, queryset):
    """
    Acción de administrador para exportar los sistemas seleccionados a un archivo Excel, incluyendo el ID.
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sistemas_exportados.xlsx"'

    libro_excel = openpyxl.Workbook()
    hoja = libro_excel.active
    hoja.title = "Sistemas"

    # Encabezados de la hoja de Excel, incluyendo ID
    hoja.append(['ID', 'Nombre', 'Sistema Principal'])

    # Iterar sobre los objetos seleccionados y agregar datos a la hoja, incluyendo ID
    for sistema in queryset:
        nombre_principal = sistema.principal.nombre if sistema.principal else ''
        hoja.append([sistema.id, sistema.nombre, nombre_principal])

    libro_excel.save(response)
    return response

exportar_sistemas_seleccionados_excel.short_description = "Exportar sistemas seleccionados a Excel (con ID)"


class HijoInline(admin.TabularInline):
    model = Sistema
    extra = 1

   

class PrincipalFilter(admin.SimpleListFilter):
    title = 'Principal'
    parameter_name = 'principal'

    def lookups(self, request, model_admin):
        # Mostrar solo los sistemas que no tienen principal en el filtro
        return [(sistema.id, sistema.nombre) for sistema in Sistema.objects.filter(principal__isnull=True)]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(principal_id=self.value())
        return queryset


@admin.register(Sistema)
class SistemaAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'principal',)
    search_fields = ('nombre',)
    list_filter = (PrincipalFilter,)
    inlines = [HijoInline]
    actions = [exportar_sistemas_seleccionados_excel]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs  # Aquí podrías aplicar filtros adicionales si lo necesitas
