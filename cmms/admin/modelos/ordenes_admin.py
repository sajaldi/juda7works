from django.contrib import admin
from ...models import OrdenDeTrabajo
from django.http import HttpResponse, HttpResponseRedirect
from openpyxl import Workbook
from django.utils import timezone

# Filtro personalizado para las fechas
class FechaOrdenFilter(admin.SimpleListFilter):
    title = 'Fecha de la orden'
    parameter_name = 'fecha'

    def lookups(self, request, model_admin):
        today = timezone.now().date()
        return (
            ('hoy', 'Hoy'),
            ('esta_semana', 'Esta semana'),
            ('este_mes', 'Este mes'),
            ('otros', 'Otros'),
        )

    def queryset(self, request, queryset):
        today = timezone.now().date()
        if self.value() == 'hoy':
            return queryset.filter(fechaDeInicio__date=today)
        elif self.value() == 'esta_semana':
            start_of_week = today - timezone.timedelta(days=today.weekday())
            return queryset.filter(fechaDeInicio__gte=start_of_week)
        elif self.value() == 'este_mes':
            return queryset.filter(fechaDeInicio__month=today.month)
        elif self.value() == 'otros':
            return queryset
        return queryset

@admin.register(OrdenDeTrabajo)
class OrdenDeTrabajoAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'HojaDeRuta', 'fechaDeInicio', 'fechaDeFin', 'area','Activo')
    list_filter = (FechaOrdenFilter,)  # Agregamos el filtro de fecha
    actions = ['exportar_plantilla', 'eliminar_intervalo', 'eliminar_sin_confirmacion']


    @admin.action(description="Eliminar sin confirmación")
    def eliminar_sin_confirmacion(self, request, queryset):
        queryset.delete()  # Elimina los registros seleccionados
        self.message_user(request, f"Se han eliminado {queryset.count()} registros correctamente.")
        
        # Redirigir a la misma página sin mostrar el listado completo
        return HttpResponseRedirect(request.get_full_path())  
    
    
    list_per_page = 800
    @admin.action(description="Exportar Plantilla de Administrador")
    def exportar_plantilla(self, request, queryset):
        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla de Administrador"

        # Definir encabezados
        headers = [
            'Nombre de la Orden de Trabajo',
            'Hoja de Ruta Asociada',
            'Fecha de Inicio',
            'Fecha de Fin',
            'Área',
            'Área Principal'  # Nueva columna
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Escribir datos de las órdenes seleccionadas
        for row_num, orden in enumerate(queryset, 2):
            area = orden.area
            nombre_area = area.nombre if area else ""
            nombre_area_principal = area.principal.nombre if area and area.principal else ""
            
            # Llenar las celdas (empieza desde la fila 2)
            ws.cell(row=row_num, column=1, value=orden.nombre)
            ws.cell(row=row_num, column=2, value=str(orden.HojaDeRuta))
            ws.cell(row=row_num, column=3, value=orden.fechaDeInicio.strftime("%Y-%m-%d %H:%M:%S") if orden.fechaDeInicio else "")
            ws.cell(row=row_num, column=4, value=orden.fechaDeFin.strftime("%Y-%m-%d %H:%M:%S") if orden.fechaDeFin else "")
            ws.cell(row=row_num, column=5, value=nombre_area)  # Área
            ws.cell(row=row_num, column=6, value=nombre_area_principal)  # Área Principal

        # Configurar la respuesta
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_admin.xlsx"'
        
        # Guardar el libro en la respuesta
        wb.save(response)
        
        return response

    @admin.action(description="Eliminar órdenes dentro del intervalo de fechas")
    def eliminar_intervalo(self, request, queryset):
        fecha_inicio = request.POST.get('fecha_inicio', None)
        fecha_fin = request.POST.get('fecha_fin', None)

        if fecha_inicio and fecha_fin:
            try:
                # Convertimos las fechas del filtro
                fecha_inicio = timezone.datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
                fecha_fin = timezone.datetime.strptime(fecha_fin, "%Y-%m-%d").date()

                # Filtrar y eliminar las órdenes dentro del intervalo
                ordenes_a_eliminar = queryset.filter(fechaDeInicio__gte=fecha_inicio, fechaDeFin__lte=fecha_fin)
                num_eliminadas = ordenes_a_eliminar.delete()

                self.message_user(request, f"{num_eliminadas[0]} órdenes eliminadas exitosamente.", level='success')
            except ValueError:
                self.message_user(request, "Formato de fechas incorrecto.", level='error')
        else:
            self.message_user(request, "Por favor, ingrese el intervalo de fechas.", level='error')
