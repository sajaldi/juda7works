# filepath: /C:/Django/SoftCoMJuda/Softcom/cmms/views.py
from pyexpat.errors import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
#from datetime import datetime, timedelta
import datetime
from openpyxl import Workbook, load_workbook
import openpyxl

from .forms import ProgramacionForm
from .models import Activo, HojaDeRuta, OrdenDeTrabajo, PasosHojaDeRuta, Sistema, Area
import logging

# Configurar el logger
logger = logging.getLogger(__name__)


from datetime import datetime, timedelta

def get_activos_por_categoria(request):
    categoria_id = request.GET.get("categoria_id")
    activos = Activo.objects.filter(modelo__categoria_id=categoria_id).values("id", "nombre")
    return JsonResponse({"activos": list(activos)})

def vista_anual(request):
    # Obtener el año actual
    year = datetime.now().year

    months = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    # Crear una lista de semanas del año
    semanas = []
    start_date = datetime(year, 1, 1)
    while start_date.weekday() != 0:  # Asegurar que empiece en lunes
        start_date += timedelta(days=1)
    for week_num in range(1, 53):
        end_date = start_date + timedelta(days=6)
        semanas.append((week_num, start_date, end_date))
        start_date = end_date + timedelta(days=1)

    # Obtener la semana actual
    fecha_actual = datetime.now()
    semana_actual = None
    for week_num, start_date, end_date in semanas:
        if start_date <= fecha_actual <= end_date:
            semana_actual = week_num
            break

    # Obtener todas las órdenes de trabajo del año
    ordenes = OrdenDeTrabajo.objects.filter(fechaDeInicio__year=year)
    logger.debug(f"Total de órdenes encontradas: {len(ordenes)}")

    # Organizar las órdenes por sistema, hoja de ruta y semana
    ordenes_por_sistema = {}
    for orden in ordenes:
        sistema = orden.HojaDeRuta.sistema
        sistema_principal = sistema.principal if sistema.principal else "Sin Sistema Principal"
        hoja_de_ruta = orden.HojaDeRuta.nombre
        # Obtener el horario de la programación en lugar de la hoja de ruta
        programacion = orden.programacion  # Asumiendo que existe una relación 'programacion' en OrdenDeTrabajo
        horario = orden.horario if programacion else None # Obtener horario de programacion
        color = orden.HojaDeRuta.intervalo.color

        if sistema_principal not in ordenes_por_sistema:
            ordenes_por_sistema[sistema_principal] = {}
        if sistema not in ordenes_por_sistema[sistema_principal]:
            ordenes_por_sistema[sistema_principal][sistema] = {}
        if hoja_de_ruta not in ordenes_por_sistema[sistema_principal][sistema]:
            ordenes_por_sistema[sistema_principal][sistema][hoja_de_ruta] = {
                'horario': horario,
                'color': color,
                'semanas': {week_num: [] for week_num in range(1, 53)}
            }
        for week_num, start_date, end_date in semanas:
            if start_date <= orden.fechaDeInicio.replace(tzinfo=None) <= end_date:
                ordenes_por_sistema[sistema_principal][sistema][hoja_de_ruta]['semanas'][week_num].append(orden)
                break

    # Calcular el primer y último nivel programado de la semana
    niveles_por_semana = {}
    for sistema_principal, sistemas in ordenes_por_sistema.items():
        for sistema, hojas_de_ruta in sistemas.items():
            for hoja_de_ruta, data in hojas_de_ruta.items():
                for week_num, ordenes in data['semanas'].items():
                    if ordenes:
                        # Aquí comprobamos si `orden.area` existe antes de acceder a `nombre`
                        niveles = [orden.area.nombre if orden.area else "Área no definida" for orden in ordenes]
                        key = f"{sistema_principal},{sistema},{hoja_de_ruta},{week_num}"
                        niveles_por_semana[key] = {
                            'niveles': f"{niveles[0]}-{niveles[-1]}",
                            'color': data['color']
                        }
                        logger.debug(f"Semana {week_num}: {niveles_por_semana[key]}")

    context = {
        'year': year,
        'semanas': semanas,
        'ordenes_por_sistema': ordenes_por_sistema,
        'niveles_por_semana': niveles_por_semana,
        'semana_actual': semana_actual,  # Pasar la semana actual al contexto
        'months': months,  # Add months to context
    }
    return render(request, 'cmms/vista_anual.html', context)


def vista_anual_filtrada(request):
    """
    Vista para mostrar un calendario anual de órdenes de trabajo, filtrado por área y sistema principal.
    """
    year = datetime.now().year
    months = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    areas = Area.objects.all()
    sistemas_principales = Sistema.objects.filter(principal__isnull=True)

    area_seleccionada_id = request.GET.get('area')
    sistema_principal_seleccionado_id = request.GET.get('sistema_principal')

    ordenes = OrdenDeTrabajo.objects.filter(fechaDeInicio__year=year)

    if area_seleccionada_id:
        ordenes = ordenes.filter(area_id=area_seleccionada_id)
    if sistema_principal_seleccionado_id:
        ordenes = ordenes.filter(HojaDeRuta__sistema__principal_id=sistema_principal_seleccionado_id)

    # *** COPIA Y ADAPTA LA LÓGICA DE vista_anual ***
    semanas = []
    start_date = datetime(year, 1, 1)
    while start_date.weekday() != 0:
        start_date += timedelta(days=1)
    for week_num in range(1, 53):
        end_date = start_date + timedelta(days=6)
        semanas.append((week_num, start_date, end_date))
        start_date = end_date + timedelta(days=1)

    fecha_actual = datetime.now()
    semana_actual = None
    for week_num, start_date, end_date in semanas:
        if start_date <= fecha_actual <= end_date:
            semana_actual = week_num
            break

    ordenes_por_sistema = {}
    for orden in ordenes:
        sistema = orden.HojaDeRuta.sistema
        sistema_principal = sistema.principal if sistema.principal else "Sin Sistema Principal"
        hoja_de_ruta = orden.HojaDeRuta.nombre
        programacion = orden.programacion
        horario = orden.horario if programacion else None
        color = orden.HojaDeRuta.intervalo.color

        if sistema_principal not in ordenes_por_sistema:
            ordenes_por_sistema[sistema_principal] = {}
        if sistema not in ordenes_por_sistema[sistema_principal]:
            ordenes_por_sistema[sistema_principal][sistema] = {}
        if hoja_de_ruta not in ordenes_por_sistema[sistema_principal][sistema]:
            ordenes_por_sistema[sistema_principal][sistema][hoja_de_ruta] = {
                'horario': horario,
                'color': color,
                'semanas': {week_num: [] for week_num in range(1, 53)}
            }
        for week_num, start_date, end_date in semanas:
            fecha_inicio_sin_tz = orden.fechaDeInicio.replace(tzinfo=None)
            if start_date <= fecha_inicio_sin_tz <= end_date:
                ordenes_por_sistema[sistema_principal][sistema][hoja_de_ruta]['semanas'][week_num].append(orden)
                break

    niveles_por_semana = {}
    for sistema_principal, sistemas in ordenes_por_sistema.items():
        for sistema, hojas_de_ruta in sistemas.items():
            for hoja_de_ruta, data in hojas_de_ruta.items():
                for week_num, ordenes in data['semanas'].items():
                    if ordenes:
                        niveles = [orden.area.nombre if orden.area else "Área no definida" for orden in ordenes]
                        key = f"{sistema_principal},{sistema},{hoja_de_ruta},{week_num}"
                        niveles_por_semana[key] = {
                            'niveles': f"{niveles[0]}-{niveles[-1]}",
                            'color': data['color']
                        }
                        logger.debug(f"Semana {week_num}: {niveles_por_semana[key]}")


    # *** FIN DE LA COPIA Y ADAPTACIÓN ***

    context = {
        'year': year,
        'semanas': semanas,
        'ordenes_por_sistema': ordenes_por_sistema,
        'niveles_por_semana': niveles_por_semana,
        'semana_actual': semana_actual,
        'areas': areas,
        'months': months,
        'sistemas_principales': sistemas_principales,
        'area_seleccionada_id': area_seleccionada_id,
        'sistema_principal_seleccionado_id': sistema_principal_seleccionado_id,
    }
    return render(request, 'cmms/vista_anual_filtrada.html', context)



def obtener_ordenes(request):
    sistema_principal = request.GET.get('sistema_principal')
    sistema = request.GET.get('sistema')
    hoja_de_ruta = request.GET.get('hoja_de_ruta')
    semana = int(request.GET.get('semana'))

    # Obtener el año actual
    year = datetime.now().year

    # Calcular el rango de fechas para la semana específica
    start_date = datetime.strptime(f'{year}-W{str(semana).zfill(2)}-1', "%Y-W%W-%w").date()
    end_date = start_date + timedelta(days=6)

    # Obtener las órdenes asociadas
    ordenes = OrdenDeTrabajo.objects.filter(
        HojaDeRuta__sistema__principal__nombre=sistema_principal,
        HojaDeRuta__sistema__nombre=sistema,
        HojaDeRuta__nombre=hoja_de_ruta,
        fechaDeInicio__range=[start_date, end_date]  # Filtrar por rango de fechas
    )

    # Preparar los datos para el modal
    datos_ordenes = [{
        'nombre': orden.nombre,
        'fechaDeInicio': orden.fechaDeInicio.strftime('%Y-%m-%d %H:%M'),
        'fechaDeFin': orden.fechaDeFin.strftime('%Y-%m-%d %H:%M') if orden.fechaDeFin else 'Sin fecha de fin',
        'area': orden.area.nombre
    } for orden in ordenes]

    return JsonResponse({'ordenes': datos_ordenes})

def vista_mensual(request, fecha_inicio):
    # Convertir la fecha de inicio a un objeto datetime
    fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
    year = fecha_inicio.year
    month = fecha_inicio.month

    # Crear una lista de semanas del mes
    semanas = []
    start_date = fecha_inicio
    while start_date.weekday() != 0:
        start_date += timedelta(days=1)
    while start_date.month == month:
        end_date = start_date + timedelta(days=6)
        semanas.append((start_date.isocalendar()[1], start_date, end_date))
        start_date = end_date + timedelta(days=1)

    # Obtener todas las órdenes de trabajo del mes
    ordenes = OrdenDeTrabajo.objects.filter(fechaDeInicio__year=year, fechaDeInicio__month=month)
    logger.debug(f"Total de órdenes encontradas: {len(ordenes)}")

    # Organizar las órdenes por sistema, hoja de ruta y semana
    ordenes_por_sistema = {}
    for orden in ordenes:
        sistema = orden.HojaDeRuta.sistema
        hoja_de_ruta = orden.HojaDeRuta.nombre
        horario = orden.horario
        if sistema not in ordenes_por_sistema:
            ordenes_por_sistema[sistema] = {}
        if hoja_de_ruta not in ordenes_por_sistema[sistema]:
            ordenes_por_sistema[sistema][hoja_de_ruta] = {'horario': horario, 'semanas': {week_num: [] for week_num, _, _ in semanas}}
        for week_num, start_date, end_date in semanas:
            if start_date <= orden.fechaDeInicio.replace(tzinfo=None) <= end_date:
                ordenes_por_sistema[sistema][hoja_de_ruta]['semanas'][week_num].append(orden)
                break

    # Calcular el primer y último nivel programado de la semana
    niveles_por_semana = {}
    for sistema, hojas_de_ruta in ordenes_por_sistema.items():
        for hoja_de_ruta, data in hojas_de_ruta.items():
            for week_num, ordenes in data['semanas'].items():
                if ordenes:
                    niveles = [orden.area.nombre for orden in ordenes]
                    key = f"{sistema},{hoja_de_ruta},{week_num}"
                    niveles_por_semana[key] = f"{niveles[0]}-{niveles[-1]}"
                    logger.debug(f"Semana {week_num}: {niveles_por_semana[key]}")

    context = {
        'year': year,
        'month': month,
        'semanas': semanas,
        'ordenes_por_sistema': ordenes_por_sistema,
        'niveles_por_semana': niveles_por_semana,
    }
    return render(request, 'cmms/vista_mensual.html', context)




def plantilla_sistema(request):
    if request.method == 'GET' and 'exportar' in request.GET:
        # Generar archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sistemas"

        # Agregar encabezados
        ws.append(["ID", "Nombre", "Sistema Principal"])

        # Agregar datos de la base de datos
        for sistema in Sistema.objects.all():
            ws.append([
                sistema.id,
                sistema.nombre,
                sistema.principal.nombre if sistema.principal else ""
            ])

        # Preparar respuesta HTTP con archivo Excel
        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=sistemas.xlsx'
        return response

    return render(request, 'cmms/plantilla_sistema.html')


from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
import json
@require_http_methods(["POST"])

def mover_ordenes_semana(request):
    try:
        data = json.loads(request.body)
        sistema_principal = data.get('sistema_principal')
        sistema_nombre = data.get('sistema')
        hoja_de_ruta_nombre = data.get('hoja_de_ruta')
        old_week = data.get('old_week')
        new_week = data.get('new_week')
        year = datetime.now().year

        # Calcular fechas de las semanas
        start_date_old = datetime.strptime(f'{year}-W{old_week:02}-1', "%Y-W%W-%w").date()
        start_date_new = datetime.strptime(f'{year}-W{new_week:02}-1', "%Y-W%W-%w").date()

        # Obtener objetos de BD
        sistema = Sistema.objects.get(principal__nombre=sistema_principal, nombre=sistema_nombre)
        hoja_de_ruta = HojaDeRuta.objects.get(sistema=sistema, nombre=hoja_de_ruta_nombre)

        # Obtener órdenes a actualizar
        ordenes = OrdenDeTrabajo.objects.filter(
            HojaDeRuta=hoja_de_ruta,
            fechaDeInicio__year=year,
            fechaDeInicio__week=old_week
        )

        # Actualizar fechas
        for orden in ordenes:
            delta_days = (orden.fechaDeInicio.date() - start_date_old).days
            nueva_fecha_inicio = start_date_new + timedelta(days=delta_days)
            orden.fechaDeInicio = nueva_fecha_inicio
            if orden.fechaDeFin:
                orden.fechaDeFin = nueva_fecha_inicio + (orden.fechaDeFin - orden.fechaDeInicio)
            orden.save()

        return JsonResponse({'status': 'success'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)}, status=400)
    

    
from datetime import datetime
from django.shortcuts import render
from .models import OrdenDeTrabajo

def vista_activos(request):
    # Obtener el año actual
    year = datetime.now().year

    # Crear la lista de semanas del año
    semanas = list(range(1, 53))

    # Obtener todas las órdenes de trabajo del año
    ordenes = OrdenDeTrabajo.objects.filter(fechaDeInicio__year=year)

    # Estructura de datos para organizar órdenes
    hoja_ruta_por_semana = {}

    for orden in ordenes:
        hoja_ruta = orden.HojaDeRuta.nombre
        activo = orden.Activo.nombre if orden.Activo else None  # Solo asignamos el activo si existe

        # Solo agregar la orden si tiene un activo
        if activo:
            # Calcular la semana de la orden
            semana_orden = orden.fechaDeInicio.isocalendar()[1]  # Obtiene el número de la semana

            # Inicializar la estructura si no existe
            if hoja_ruta not in hoja_ruta_por_semana:
                hoja_ruta_por_semana[hoja_ruta] = {semana: set() for semana in semanas}

            # Agregar el activo a la semana correspondiente (solo si no está ya presente)
            hoja_ruta_por_semana[hoja_ruta][semana_orden].add(activo)

    # Convertir a formato más accesible para el template
    semanas_data = {}
    for hoja_ruta, semanas_info in hoja_ruta_por_semana.items():
        semanas_data[hoja_ruta] = [list(semanas_info.get(semana, [])) for semana in semanas]

    context = {
        'semanas': semanas,
        'hoja_ruta_por_semana': semanas_data
    }

    return render(request, 'cmms/vista_activos.html', context)




from datetime import datetime, timedelta
from django.shortcuts import render
from .models import OrdenDeTrabajo

def vista_activos_por_dia(request, fecha_inicio):
    # Convertir la fecha desde la URL a un objeto datetime
    fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    year = fecha_inicio.year
    semana_seleccionada = fecha_inicio.isocalendar()[1]

    # Calcular las fechas de los días de la semana seleccionada
    primer_dia_semana = fecha_inicio - timedelta(days=fecha_inicio.weekday())
    dias_semana = [(primer_dia_semana + timedelta(days=i)) for i in range(7)]

    # Obtener todas las órdenes de trabajo de la semana seleccionada
    ordenes = OrdenDeTrabajo.objects.filter(fechaDeInicio__year=year, fechaDeInicio__week=semana_seleccionada)

    # Estructura de datos para organizar órdenes por día
    activos_por_dia = {dia: set() for dia in dias_semana}

    for orden in ordenes:
        fecha_orden = orden.fechaDeInicio.date()
        activo = orden.Activo.nombre if orden.Activo else None

        if activo and fecha_orden in activos_por_dia:
            activos_por_dia[fecha_orden].add(activo)

    # Convertir los sets en listas para facilitar la iteración en el template
    activos_por_dia = {dia: list(activos) for dia, activos in activos_por_dia.items()}

    context = {
        'dias_semana': dias_semana,
        'activos_por_dia': activos_por_dia,
        'semana_actual': semana_seleccionada,
    }

    return render(request, 'cmms/vista_activos_por_dia.html', context)


def exportar_pasos_hoja_de_ruta(request):
    # Crear un nuevo libro de trabajo y hoja
     # Crear un nuevo libro de trabajo y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pasos de Hoja de Ruta"

    # Encabezados con formato en negrita
    encabezados = ["ID", "Paso", "Tiempo (min)", "Hoja de Ruta",  "Sistema Principal", "Sistema","ordenamiento","frecuencia"]


    ws.append(encabezados)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Obtener datos optimizando la consulta
    for paso in PasosHojaDeRuta.objects.select_related('hojaderuta').all():
        sistema_principal = paso.hojaderuta.sistema.principal.nombre if paso.hojaderuta.sistema and paso.hojaderuta.sistema.principal else "N/A"
        sistema = paso.hojaderuta.sistema.nombre if paso.hojaderuta.sistema else "N/A"
        ordenamiento = paso.hojaderuta.ordenamiento if paso.hojaderuta.ordenamiento else 0
        frecuencia = str(paso.hojaderuta.intervalo) if paso.hojaderuta.intervalo else "N/A"


        ws.append([paso.id, paso.paso, paso.tiempo, paso.hojaderuta.nombre, sistema_principal, sistema, ordenamiento,frecuencia])

    # Ajustar ancho de columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Configurar respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="PasosHojaDeRuta.xlsx"'
    
    try:
        wb.save(response)
    except Exception as e:
        return HttpResponse(f"Error al generar el archivo: {e}")

    return response

import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import HojaDeRuta, PasosHojaDeRuta


import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import PasosHojaDeRuta, HojaDeRuta

def importar_pasos_hoja_de_ruta(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Error al abrir el archivo: {str(e)}")
            return redirect("importar_pasos")

        errores = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # Saltamos encabezados
            print(f"Fila {i}: {row}")  # 🔍 Depuración: ver qué datos está leyendo

            # Tomar solo las primeras 4 columnas (ID, Paso, Tiempo, Hoja de Ruta)
            datos = row[:4]  # Esto ignora las columnas extra
            if len(datos) < 4:
                errores.append(f"Fila {i}: Número incorrecto de columnas ({len(datos)}). Se esperaban al menos 4 (ID, Paso, Tiempo, Hoja de Ruta).")
                continue

            paso_id, paso_nombre, tiempo, hoja_de_ruta_nombre = datos

            # Si el paso a importar tiene la palabra "ELIMINAR", intentamos eliminarlo
            if str(paso_nombre).strip().upper() == "ELIMINAR":
                try:
                    eliminado = PasosHojaDeRuta.objects.filter(id=paso_id).delete()
                    if eliminado[0] > 0:
                        print(f"Fila {i}: Paso con ID '{paso_id}' eliminado.")
                    else:
                        errores.append(f"Fila {i}: No se encontró el paso con ID '{paso_id}' para eliminar.")
                except Exception as e:
                    errores.append(f"Fila {i}: Error al eliminar el paso con ID '{paso_id}' ({str(e)}).")
                continue  # Pasamos a la siguiente fila sin intentar crear/actualizar

            # Validar que los valores Paso Nombre, Tiempo y Hoja de Ruta Nombre no sean None
            if paso_nombre is None or tiempo is None or hoja_de_ruta_nombre is None:
                errores.append(f"Fila {i}: Hay celdas vacías en la fila (Paso, Tiempo, Hoja de Ruta Nombre).")
                continue

            # Validar que el tiempo sea numérico
            if not isinstance(tiempo, (int, float)):
                errores.append(f"Fila {i}: El tiempo '{tiempo}' no es un número válido.")
                continue

            # Validar existencia de la hoja de ruta
            hoja = HojaDeRuta.objects.filter(nombre=hoja_de_ruta_nombre).first()
            if not hoja:
                errores.append(f"Fila {i}: La hoja de ruta '{hoja_de_ruta_nombre}' no existe.")
                continue

            # Intentar actualizar o crear el paso basado en el ID
            try:
                paso_obj, created = PasosHojaDeRuta.objects.update_or_create(
                    id=paso_id,  # Usamos el ID del archivo para buscar o crear
                    defaults={  # Campos a actualizar o crear si no existe
                        'paso': paso_nombre,
                        'tiempo': tiempo,
                        'hojaderuta': hoja
                    }
                )
                if created:
                    print(f"Fila {i}: Paso con ID '{paso_id}' creado.")
                else:
                    print(f"Fila {i}: Paso con ID '{paso_id}' actualizado.")
            except Exception as e:
                errores.append(f"Fila {i}: Error al guardar/actualizar en la base de datos para ID '{paso_id}' ({str(e)}).")

        # Mostrar mensajes al usuario
        if errores:
            messages.error(request, "Errores en la importación:\n" + "\n".join(errores))
        else:
            messages.success(request, "Importación exitosa.")

        return redirect("importar_pasos")  # Ajusta según tu vista de redirección

    return render(request, "cmms/importar_pasos.html")



from django.http import HttpResponse

def menu(request):
    # Aquí puedes poner lo que quieras que devuelva la vista
    # En este caso, solo retornamos un mensaje simple
    return render(request, 'cmms/menu.html')  # Renderiza la plantilla



from datetime import datetime, timedelta

def ordenes_por_fecha(request, semana):
    """
    Vista que, dado el número de semana, muestra una tabla con las fechas de esa semana
    como encabezado y los activos programados por fecha.
    """
    try:
        semana = int(semana)
        if semana < 1 or semana > 52:
            raise ValueError("Número de semana fuera de rango")
    except ValueError:
        semana = 1  # Valor por defecto en caso de error

    year = 2025

    try:
        monday = datetime.strptime(f'{year}-W{semana-1}-1', "%Y-W%U-%w").date()
    except Exception as e:
        monday = datetime(year, 1, 1).date()  # Fallback si la conversión falla

    fechas = [monday + timedelta(days=i) for i in range(7)]

    # Ajustar la consulta para incluir órdenes que tengan fechaDeInicio dentro del rango de la semana
    orders = OrdenDeTrabajo.objects.filter(fechaDeInicio__gte=fechas[0], fechaDeInicio__lte=fechas[-1])

    rutas = orders.values_list('HojaDeRuta', flat=True).distinct()

    hoja_ruta_por_fecha = {}
    for ruta in rutas:
        hoja_ruta_por_fecha[ruta] = []
        for fecha in fechas:
            # Ahora se consulta todas las órdenes para ese día en particular, no solo las que tienen la fecha exacta
            activos = orders.filter(HojaDeRuta=ruta, fechaDeInicio__date=fecha).values_list('Activo__nombre', flat=True)
            hoja_ruta_por_fecha[ruta].append(list(activos))

    context = {
         'semana': semana,
         'fechas': fechas,
         'hoja_ruta_por_fecha': hoja_ruta_por_fecha,
    }
    return render(request, 'cmms/ordenes_por_fecha.html', context)



from django.shortcuts import render

def menu_view(request):
    # Aquí podrías hacer una consulta a la base de datos para obtener los elementos dinámicos
    dynamic_menu_items = [
        {'name': 'Nuevo Item 1', 'url': '/nuevo_item_1/'},
        {'name': 'Nuevo Item 2', 'url': '/nuevo_item_2/'},
        # Otros elementos dinámicos
    ]
    
    return render(request, 'nombre_de_tu_template.html', {'dynamic_menu_items': dynamic_menu_items})


from .models import MenuItem

def menu_view(request):
    dynamic_menu_items = MenuItem.objects.all().order_by('order')
    return render(request, 'nombre_de_tu_template.html', {'dynamic_menu_items': dynamic_menu_items})



def programacion_view(request):
    if request.method == 'POST':
        form = ProgramacionForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirigir o hacer alguna acción después de guardar
    else:
        form = ProgramacionForm()

    return render(request, 'cmms/programacion_form.html', {'form': form})

from .forms import ImportForm, ExportForm
from .utils import import_from_excel, export_to_excel  # Asegúrate de que estas funciones estén importadas

def import_view(request):
    if request.method == 'POST' and request.FILES['file']:
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            model_name = form.cleaned_data['model']
            file = request.FILES['file']
            file_path = f"media/{file.name}"
            
            # Guardamos el archivo en el sistema de archivos
            with open(file_path, 'wb') as f:
                for chunk in file.chunks():
                    f.write(chunk)
            
            try:
                message = import_from_excel(model_name, file_path)
                messages.success(request, message)
            except Exception as e:
                messages.error(request, f"Error al importar: {str(e)}")
    else:
        form = ImportForm()

    return render(request, 'import.html', {'form': form})

def export_view(request):
    if request.method == 'POST':
        form = ExportForm(request.POST)
        if form.is_valid():
            model_name = form.cleaned_data['model']
            file_path = export_to_excel(model_name)
            with open(file_path, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={model_name}_export.xlsx'
                return response
    else:
        form = ExportForm()

    return render(request, 'export.html', {'form': form})
