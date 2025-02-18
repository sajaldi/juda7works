from pyexpat.errors import messages
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from cmms.models import HojaDeRuta, Frecuencia, Sistema


def exportar_plantilla_hojaderuta(request):
    # Crear un libro de trabajo y una hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla de Exportación de HojaDeRuta"

    # Agregar encabezados
    headers = ["ID","clave_rutina", "Nombre", "Descripción", "Intervalo", "Sistema", "Ordenamiento"]
    ws.append(headers)

    # Obtener los datos del modelo HojaDeRuta
    hojas_de_ruta = HojaDeRuta.objects.all()
    for hoja in hojas_de_ruta:
        intervalo_nombre = hoja.intervalo.nombre if hoja.intervalo else ''
        sistema_nombre = hoja.sistema.nombre if hoja.sistema else ''
        ws.append([hoja.id, hoja.clave_rutina, hoja.nombre, hoja.descripcion, intervalo_nombre, sistema_nombre, hoja.ordenamiento])

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_exportacion_hojaderuta.xlsx'

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)
    return response


def importar_plantilla_hojaderuta(request):
    if request.method == 'POST':
        uploaded_file = request.FILES['file']
        wb = load_workbook(uploaded_file)
        ws = wb.active

        errores = []
        importados = 0

        # Leer los datos del archivo Excel
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                hoja_id, clave_rutina, nombre, descripcion, intervalo_nombre, sistema_nombre, hoja_ordenamiento = row

                # Verificar si hoja_id es None
                if hoja_id is None:
                    errores.append(f"Error en la fila {row}: El ID de Hoja de Ruta no puede estar vacío.")
                    continue # Saltar esta fila si hoja_id es None

                # Obtener o crear el intervalo
                intervalo = None # Inicializar a None por defecto
                if intervalo_nombre: # Verificar que intervalo_nombre no esté vacío
                    intervalo, created = Frecuencia.objects.get_or_create(nombre=intervalo_nombre.strip()) # Añadido strip() para eliminar espacios en blanco
                    if created:
                        print(f"Nueva frecuencia creada: {intervalo_nombre}")

                # Obtener o crear el sistema
                sistema = None # Inicializar a None por defecto
                if sistema_nombre: # Verificar que sistema_nombre no esté vacío
                    sistema, created = Sistema.objects.get_or_create(nombre=sistema_nombre.strip()) # Añadido strip() para eliminar espacios en blanco
                    if created:
                        print(f"Nuevo sistema creado: {sistema_nombre}")

                # Actualizar o crear la hoja de ruta
                hoja, created = HojaDeRuta.objects.update_or_create(
                    id=hoja_id, # Usa el ID del Excel para buscar o crear
                    defaults={
                        'clave_rutina': clave_rutina.strip() if clave_rutina else None, # Añadido strip() y manejo de None
                        'nombre': nombre.strip() if nombre else None, # Añadido strip() y manejo de None
                        'descripcion': descripcion.strip() if descripcion else None, # Añadido strip() y manejo de None
                        'intervalo': intervalo,
                        'sistema': sistema,
                        'ordenamiento': hoja_ordenamiento.strip() if isinstance(hoja_ordenamiento, str) else hoja_ordenamiento, # Solo aplica strip() si es una cadena
                    }
                )
                importados += 1

            except Exception as e:
                errores.append(f"Error en la fila {row}: {str(e)}")

        context = {
            'errores': errores,
            'importados': importados,
            'mensaje': f'Se importaron {importados} hojas de ruta exitosamente.' if importados > 0 else None
        }
        return render(request, 'cmms/importar_hojaderuta.html', context)

    return render(request, 'cmms/importar_hojaderuta.html')